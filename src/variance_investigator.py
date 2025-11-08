import pandas as pd
import openpyxl
from openpyxl import load_workbook
from pathlib import Path
import os
import openai
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Configuration
EXCEL_FILE = "AP_Reconciliation.xlsx"
EMAILS_FILE = "Data/emails/demo_inbox_emails.xlsx"
THRESHOLD_PERCENT = 10.0
AP_ACCOUNT_CODE = "2000"  # Accounts Payable account code
OPENAI_API_KEY = os.getenv('OPENAI_API_KEY')

def read_mom_percentage():
    """Read the MoM % Change from the Excel file"""
    try:
        # First try to get calculated values
        wb = load_workbook(EXCEL_FILE, data_only=True)
        ws = wb["AP Reconciliation Summary"]

        # Find the MoM % Change cell
        for row in range(1, 20):
            cell_a = ws[f"A{row}"]
            if cell_a.value and "MoM % Change" in str(cell_a.value):
                cell_b = ws[f"B{row}"]
                mom_percent = cell_b.value
                print(f"Found MoM % Change (calculated): {mom_percent}")

                if mom_percent is not None:
                    # Handle percentage formatting
                    if isinstance(mom_percent, str):
                        mom_percent = mom_percent.replace('%', '').strip()
                    return float(mom_percent)

        # If no calculated value, try to calculate manually
        wb = load_workbook(EXCEL_FILE, data_only=False)
        ws = wb["AP Reconciliation Summary"]

        # Get GL Balance August (row 3, column B)
        aug_balance = None
        sep_balance = None

        for row in range(1, 20):
            cell_a = ws[f"A{row}"]
            cell_b = ws[f"B{row}"]
            if cell_a.value == "GL Balance August":
                aug_balance = float(cell_b.value) if cell_b.value else 0.0
            elif cell_a.value == "GL Balance September":
                sep_balance = float(cell_b.value) if cell_b.value else 0.0

        if aug_balance is not None and sep_balance is not None and aug_balance != 0:
            mom_percent = ((sep_balance - aug_balance) / aug_balance) * 100
            print(f"Calculated MoM % Change: {mom_percent:.2f}%")
            return mom_percent

        print("Could not calculate MoM percentage")
        return 0.0

    except Exception as e:
        print(f"Error reading MoM percentage: {e}")
        return 0.0

def read_ap_summary():
    """Read all information from AP Reconciliation Summary tab"""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb["AP Reconciliation Summary"]

        # Extract all content from the summary sheet
        summary_info = []
        for row in range(1, 50):  # Check first 50 rows
            row_data = []
            for col in range(1, 10):  # Check first 10 columns
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    row_data.append(f"{chr(64+col)}{row}: {cell.value}")  # Add cell reference like A1: Value
            if row_data:
                summary_info.append(" | ".join(row_data))

        summary_text = "\n".join(summary_info)
        print(f"AP Summary content preview: {summary_text[:500]}...")
        return summary_text

    except Exception as e:
        print(f"Error reading AP summary: {e}")
        return ""

def read_reconciliation_detail():
    """Read information from Reconciliation Detail tab including images"""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Reconciliation Detail"]

        # Extract any text content from the detail sheet
        detail_info = []
        for row in range(1, 50):  # Check first 50 rows
            for col in range(1, 10):  # Check first 10 columns
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    detail_info.append(str(cell.value))

        detail_text = " ".join(detail_info)
        print(f"Reconciliation Detail content: {detail_text[:500]}...")

        # Check for images in the worksheet
        image_info = ""
        try:
            # Check for image indicators in the text
            if "screenshot" in detail_text.lower() or "invoice" in detail_text.lower():
                image_info = " Contains supporting invoice screenshots or document images for review."
        except:
            image_info = " May contain supporting documentation images."

        return detail_text + image_info

    except Exception as e:
        print(f"Error reading reconciliation detail: {e}")
        return ""

def read_emails_for_investigation():
    """Read emails from demo_inbox_emails.xlsx for investigation context"""
    try:
        df = pd.read_excel(EMAILS_FILE)
        print(f"Loaded {len(df)} emails from {EMAILS_FILE}")

        # Get all email content as structured text
        email_details = []
        for _, row in df.iterrows():
            email_info = []
            for col in df.columns:
                if pd.notnull(row[col]):
                    email_info.append(f"{col}: {row[col]}")
            email_details.append(" | ".join(email_info))

        combined_emails = "\n".join(email_details)
        print(f"Email content preview: {combined_emails[:500]}...")
        return combined_emails

    except Exception as e:
        print(f"Error reading emails: {e}")
        return ""

def read_september_gl_entries():
    """Read GL entries for September month and AP account code 2000"""
    try:
        gl_file = "Data/gl/sample_gl_aug_sep_cleaned.xlsx"
        df = pd.read_excel(gl_file)
        print(f"Loaded {len(df)} GL entries from {gl_file}")

        # Filter for September entries (month 9) and AP account code
        df['date'] = pd.to_datetime(df['date'], errors='coerce')
        sept_entries = df[df['date'].dt.month == 9]

        # Further filter for AP account code 2000 (assuming 'account' column)
        if 'account' in df.columns:
            sept_ap_entries = sept_entries[sept_entries['account'].astype(str) == AP_ACCOUNT_CODE]
        else:
            # Fallback: try 'account_code' or 'acct_code' or similar
            account_col = None
            for col in df.columns:
                if 'account' in col.lower():
                    account_col = col
                    break
            if account_col:
                sept_ap_entries = sept_entries[sept_entries[account_col].astype(str) == AP_ACCOUNT_CODE]
            else:
                print("Warning: Could not find account column in GL file, using all September entries")
                sept_ap_entries = sept_entries

        print(f"Found {len(sept_ap_entries)} September AP GL entries for account {AP_ACCOUNT_CODE}")

        # Format GL entries for AI analysis
        gl_details = []
        for _, row in sept_ap_entries.iterrows():
            gl_info = []
            for col in df.columns:
                if pd.notnull(row[col]):
                    gl_info.append(f"{col}: {row[col]}")
            gl_details.append(" | ".join(gl_info))

        combined_gl = "\n".join(gl_details)
        print(f"September AP GL entries preview: {combined_gl[:500]}...")
        return combined_gl

    except Exception as e:
        print(f"Error reading September GL entries: {e}")
        return ""

def generate_investigation_reason(mom_percent, summary_info, detail_info, sept_gl_info, email_content):
    """Use OpenAI API to generate a reason for the large movement"""

    if not OPENAI_API_KEY:
        return "Error: OpenAI API key not found in environment variables"

    # Prepare the context for OpenAI
    direction = "increase" if mom_percent > 0 else "decrease"

    # No GL context needed as requested

    prompt = f"""
    ROLE: You are a Senior Accounts Payable Accountant conducting a variance investigation.

    ANALYSIS CONTEXT:
    - Reviewing Accounts Payable account code {AP_ACCOUNT_CODE}
    - Significant variance detected: {mom_percent:.1f}% change month-over-month ({direction})
    - AP balances increase when invoices are received, decrease when payments are made

    SUPPORTING INFORMATION FROM EXCEL RECONCILIATION FILE:

    AP RECONCILIATION SUMMARY TAB:
    {summary_info}

    RECONCILIATION DETAIL TAB:
    {detail_info}

    SEPTEMBER GENERAL LEDGER ENTRIES:
    {sept_gl_info}

    EMAIL COMMUNICATIONS DATA:
    {email_content[:1500]}  # Limited to prevent token overflow

    INVESTIGATION TASK:
    Based on ALL the information above from the Excel reconciliation file and email communications, provide a comprehensive conclusion explaining why the AP balance moved significantly.

    Your conclusion should:
    1. Analyze the specific data from the Reconciliation Detail tab
    2. Correlate findings with the email communications
    3. Explain the business reason for the movement
    4. Identify any red flags or areas requiring further investigation
    5. Be written as a professional accounting analysis suitable for financial reporting
    """

    try:
        client = openai.OpenAI(api_key=OPENAI_API_KEY)

        response = client.chat.completions.create(
            model="gpt-4",
            messages=[
                {"role": "system", "content": "You are a Senior Accounts Payable Accountant with extensive experience in financial reconciliation, variance analysis, and audit documentation. You write clear, professional explanations that are suitable for financial reporting and audit reviews. Your explanations are evidence-based and include appropriate recommendations for further investigation."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=400,
            temperature=0.2  # Lower temperature for more professional, consistent responses
        )

        content = response.choices[0].message.content
        if content:
            reason = content.strip()
        else:
            reason = "No analysis could be generated from the provided data"
        return reason

    except Exception as e:
        print(f"Error calling OpenAI API: {e}")
        return f"Unable to generate AI analysis due to API error: {str(e)}"

def clear_previous_reason():
    """Clear any existing 'Reason' from the Excel summary sheet"""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb["AP Reconciliation Summary"]

        # Find and remove the Reason row
        for row in range(1, 50):
            if ws[f"A{row}"].value and str(ws[f"A{row}"].value).lower() == "reason":
                # Clear the row
                ws[f"A{row}"] = None
                ws[f"B{row}"] = None
                print(f"Cleared existing reason from Excel")
                break

        wb.save(EXCEL_FILE)

    except Exception as e:
        print(f"Error clearing previous reason: {e}")

def update_excel_with_reason(reason):
    """Add or update a 'Reason' line in the Excel summary sheet"""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb["AP Reconciliation Summary"]

        # Check if Reason already exists
        reason_row = None
        for row in range(1, 50):
            if ws[f"A{row}"].value and str(ws[f"A{row}"].value).lower() == "reason":
                reason_row = row
                break

        if reason_row:
            # Update existing reason
            ws[f"B{reason_row}"] = reason
            print(f"Updated existing reason in Excel")
        else:
            # Find the last row with data and add reason
            last_row = 1
            for row in range(1, 50):
                if ws[f"A{row}"].value:
                    last_row = row

            # Add reason in the next row
            ws[f"A{last_row + 1}"] = "Reason"
            ws[f"B{last_row + 1}"] = reason
            print(f"Added new reason to Excel")

        wb.save(EXCEL_FILE)

    except Exception as e:
        print(f"Error updating Excel with reason: {e}")

def main():
    print("🔍 Starting Variance Investigation...")

    # Clear previous reason from Excel
    clear_previous_reason()

    # Read MoM percentage
    mom_percent = read_mom_percentage()

    # Check if investigation is needed
    if abs(mom_percent) <= THRESHOLD_PERCENT:
        print(f"MoM % Change ({mom_percent:.1f}%) is within threshold ({THRESHOLD_PERCENT}%), no investigation needed.")
        return

    print(f"MoM % Change ({mom_percent:.1f}%) exceeds threshold, investigating...")

    # Gather investigation data
    summary_info = read_ap_summary()
    detail_info = read_reconciliation_detail()
    sept_gl_info = read_september_gl_entries()
    email_content = read_emails_for_investigation()

    # Generate AI-based reason
    reason = generate_investigation_reason(mom_percent, summary_info, detail_info, sept_gl_info, email_content)

    # Update Excel with reason
    update_excel_with_reason(reason)

    print("✅ Variance investigation complete.")
    print(f"Reason: {reason}")

if __name__ == "__main__":
    main()